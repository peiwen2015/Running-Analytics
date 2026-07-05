#!/usr/bin/env python3
from __future__ import annotations

import html
import json
import mimetypes
import socket
import os
import sys
import subprocess
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse

from openpyxl import load_workbook

from fit_to_excel import (
    APP_VERSION,
    DEFAULT_OUTPUT_DIR,
    DROPDOWN_CONFIG_PATH,
    WORKBOOK_VERSION_NAME,
    create_workbook,
    load_dropdown_options,
)


ROOT = Path(__file__).resolve().parent
FIT_DIR = ROOT / "FIT"
ASSETS_DIR = ROOT / "assets"
HOST = "127.0.0.1"
PORT = 8765
EXCEL_FORMAT_VERSION = WORKBOOK_VERSION_NAME
EXCEL_SCHEMA_LABEL = EXCEL_FORMAT_VERSION.replace("跑步分析資料 ", "Excel Schema ")
DEFAULT_FIT_LIST_LIMIT = 30
OPTION_FIELDS = [
    ("shoes", "鞋款"),
    ("workout_types", "課表類型"),
    ("training_focus", "訓練目的"),
    ("garmin_rpe", "Garmin 主觀感受"),
]
WORKOUT_FOCUS_MAP_KEY = "workout_focus_map"
DEFAULT_WORKOUT_FOCUS_HINTS = {
    "Recovery": ["Recovery"],
    "Easy": ["Aerobic Base"],
    "LSD": ["Endurance"],
    "Long Run": ["Endurance"],
    "Progression": ["Endurance"],
    "Tempo": ["Threshold"],
    "Marathon Pace": ["Marathon Pace"],
    "Interval": ["VO₂max"],
    "Repetition": ["Speed"],
    "Hill": ["Running Economy"],
    "Race": ["Race Simulation", "Test"],
    "Test": ["Race Simulation", "Test"],
    "Fartlek": ["Speed", "VO₂max"],
}


def open_file(path):
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.run(["open", str(path)], check=False)
    else:
        subprocess.run(["xdg-open", str(path)], check=False)


def is_output_file(path):
    try:
        resolved = path.resolve()
        output_root = DEFAULT_OUTPUT_DIR.resolve()
        return resolved.is_file() and resolved.suffix.lower() == ".xlsx" and resolved.is_relative_to(output_root)
    except OSError:
        return False


def pace_text(seconds, meters):
    if not seconds or not meters:
        return ""
    sec_per_km = round(float(seconds) / float(meters) * 1000)
    return f"{sec_per_km // 60}:{sec_per_km % 60:02d}/km"


def format_duration(seconds):
    if not seconds:
        return ""
    seconds = int(round(float(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def cell_map(ws):
    return {ws.cell(row, 1).value: ws.cell(row, 2).value for row in range(1, ws.max_row + 1)}


def load_raw_config():
    if not DROPDOWN_CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(DROPDOWN_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def matching_option(options, hint):
    hint = hint.lower()
    for option in options:
        if hint in str(option).lower():
            return option
    return None


def default_workout_focus_map(options):
    result = {}
    training_focus = options.get("training_focus", [])
    for workout in options.get("workout_types", []):
        matched_focus = []
        for workout_hint, focus_hints in DEFAULT_WORKOUT_FOCUS_HINTS.items():
            if workout_hint.lower() not in str(workout).lower():
                continue
            for focus_hint in focus_hints:
                focus = matching_option(training_focus, focus_hint)
                if focus and focus not in matched_focus:
                    matched_focus.append(focus)
            break
        result[workout] = matched_focus
    return result


def clean_workout_focus_map(raw_map, options):
    workout_types = options.get("workout_types", [])
    training_focus = set(options.get("training_focus", []))
    result = {}
    if isinstance(raw_map, dict):
        for workout in workout_types:
            values = raw_map.get(workout, [])
            if isinstance(values, str):
                values = [values]
            if isinstance(values, list):
                result[workout] = [
                    str(value).strip()
                    for value in values
                    if str(value).strip() in training_focus
                ]
    defaults = default_workout_focus_map(options)
    for workout in workout_types:
        result.setdefault(workout, defaults.get(workout, []))
    return result


def load_app_options():
    options = load_dropdown_options(DROPDOWN_CONFIG_PATH)
    raw = load_raw_config()
    options[WORKOUT_FOCUS_MAP_KEY] = clean_workout_focus_map(raw.get(WORKOUT_FOCUS_MAP_KEY), options)
    return options


def workbook_summary(path):
    wb = load_workbook(path, data_only=True)
    info = cell_map(wb["活動資訊"])
    km = wb["每公里數據"]
    rows = [
        row
        for row in km.iter_rows(min_row=3, max_row=km.max_row, values_only=True)
        if row and isinstance(row[0], int)
    ]
    total_distance = sum(row[1] or 0 for row in rows)
    total_seconds = sum(row[2] or 0 for row in rows)
    avg_hr_values = [row[4] for row in rows if isinstance(row[4], (int, float))]
    avg_power_values = [row[8] for row in rows if isinstance(row[8], (int, float))]

    summary = [
        ("活動日期", info.get("活動日期")),
        ("開始時間", info.get("開始時間")),
        ("距離", f"{total_distance / 1000:.2f} km" if total_distance else ""),
        ("時間", format_duration(total_seconds)),
        ("平均配速", pace_text(total_seconds, total_distance)),
        ("平均心率", round(sum(avg_hr_values) / len(avg_hr_values), 1) if avg_hr_values else ""),
        ("平均功率", f"{round(sum(avg_power_values) / len(avg_power_values), 1)} W" if avg_power_values else ""),
        ("天氣", weather_summary(info)),
        ("Training Effect", training_effect_summary(info)),
        ("Training Load", info.get("Training Load")),
    ]
    return [(label, value) for label, value in summary if value not in ("", None)]


def weather_summary(info):
    temp = info.get("天氣氣溫 (°C)", info.get("天氣氣溫(°C)"))
    humidity = info.get("濕度 (%)", info.get("濕度(%)"))
    wind_direction = info.get("風向")
    wind_speed = info.get("風速")
    parts = []
    if temp not in ("", None):
        parts.append(f"{temp}°C")
    if humidity not in ("", None):
        parts.append(f"{humidity}%")
    wind = " ".join(str(value) for value in (wind_direction, wind_speed) if value not in ("", None))
    if wind:
        parts.append(wind)
    return " / ".join(parts)


def training_effect_summary(info):
    aerobic = info.get("Training Effect (Aerobic)")
    anaerobic = info.get("Training Effect (Anaerobic)")
    parts = []
    if aerobic not in ("", None):
        parts.append(f"Aerobic {aerobic}")
    if anaerobic not in ("", None):
        parts.append(f"Anaerobic {anaerobic}")
    return " / ".join(parts)


def summary_html(items):
    if not items:
        return ""
    rows = "\n".join(
        f"<tr><th>{html.escape(str(label))}</th><td>{html.escape(str(value))}</td></tr>"
        for label, value in items
    )
    return f"""
      <table class="summary">
        <tbody>
          {rows}
        </tbody>
      </table>
    """


def friendly_error(error):
    text = str(error)
    if isinstance(error, FileNotFoundError):
        return "找不到指定的檔案，請確認 FIT 檔還在原本的位置。"
    if isinstance(error, PermissionError):
        return "目前沒有權限讀寫這個檔案或資料夾，請確認檔案沒有被 Excel 開著，或換一個輸出檔名再試一次。"
    if isinstance(error, ModuleNotFoundError):
        return "缺少必要套件，請重新啟動應用程式，讓啟動檔自動安裝需求套件。"
    if isinstance(error, socket.timeout):
        return "天氣查詢逾時。可以稍後再試，或先取消自動抓天氣完成轉檔。"
    if "No lap data found" in text:
        return "這個 FIT 裡沒有可用的每公里分段資料，可能不是跑步活動，或檔案內容不完整。"
    if "FIT decode errors" in text:
        return "FIT 檔解析失敗，請確認這是 Garmin Connect 匯出的 Original FIT 檔。"
    if "urlopen" in text or "Open-Meteo" in text:
        return "天氣查詢失敗。請確認網路可用，或先取消自動抓天氣完成轉檔。"
    return f"轉檔失敗：{text}"


def all_fit_files():
    FIT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(FIT_DIR.glob("*.fit"), key=lambda path: path.stat().st_mtime, reverse=True)


def fit_files(selected_fit=""):
    files = all_fit_files()
    total_count = len(files)
    limit = DEFAULT_FIT_LIST_LIMIT
    limited = files[:limit]
    if selected_fit and not any(path.name == selected_fit for path in limited):
        selected_path = FIT_DIR / selected_fit
        if selected_path.exists() and selected_path.suffix.lower() == ".fit":
            limited.insert(0, selected_path)
    return limited, total_count, limit


def parse_number(value):
    value = (value or "").strip()
    if not value:
        return ""
    try:
        return float(value)
    except ValueError:
        return value


def first_value(form, key, default=""):
    return (form.get(key, [default])[0] or "").strip()


def selected_values(form, key):
    return [value.strip() for value in form.get(key, []) if value.strip()]


def content_disposition_params(value):
    params = {}
    for part in value.split(";"):
        part = part.strip()
        if "=" not in part:
            continue
        key, raw = part.split("=", 1)
        params[key.lower()] = raw.strip().strip('"')
    return params


def parse_multipart(body, content_type):
    marker = "boundary="
    if marker not in content_type:
        return {}, {}
    boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"').encode()
    delimiter = b"--" + boundary
    form = {}
    files = {}

    for part in body.split(delimiter):
        if not part or part in (b"--\r\n", b"--"):
            continue
        if part.startswith(b"\r\n"):
            part = part[2:]
        if part.endswith(b"--\r\n"):
            part = part[:-4]
        elif part.endswith(b"--"):
            part = part[:-2]
        if part.endswith(b"\r\n"):
            part = part[:-2]
        if b"\r\n\r\n" not in part:
            continue

        raw_headers, data = part.split(b"\r\n\r\n", 1)
        headers = {}
        for line in raw_headers.decode("utf-8", "replace").split("\r\n"):
            if ":" in line:
                key, value = line.split(":", 1)
                headers[key.lower()] = value.strip()
        disposition = headers.get("content-disposition", "")
        params = content_disposition_params(disposition)
        name = params.get("name")
        if not name:
            continue
        filename = params.get("filename")
        if filename:
            files[name] = {"filename": Path(filename).name, "content": data}
        else:
            form.setdefault(name, []).append(data.decode("utf-8", "replace"))
    return form, files


def parse_post_data(headers, body):
    content_type = headers.get("Content-Type", "")
    if content_type.startswith("multipart/form-data"):
        return parse_multipart(body, content_type)
    return parse_qs(body.decode("utf-8")), {}


def save_uploaded_fit(upload):
    if not upload:
        return None
    filename = upload.get("filename") or ""
    content = upload.get("content") or b""
    if not filename or not content:
        return None
    if Path(filename).suffix.lower() != ".fit":
        raise ValueError("上傳檔案必須是 .fit。")
    FIT_DIR.mkdir(parents=True, exist_ok=True)
    target = FIT_DIR / Path(filename).name
    target.write_bytes(content)
    return target


def build_metadata(form):
    return {
        "shoe": first_value(form, "shoe"),
        "weather_temp": parse_number(first_value(form, "weather_temp")),
        "humidity": parse_number(first_value(form, "humidity")),
        "wind_direction": first_value(form, "wind_direction"),
        "wind_speed": first_value(form, "wind_speed"),
        "workout_type": first_value(form, "workout_type"),
        "training_focus": "、".join(selected_values(form, "training_focus")),
        "rpe": first_value(form, "rpe"),
        "fueling": first_value(form, "fueling"),
        "max_hr": parse_number(first_value(form, "max_hr")),
        "critical_power": parse_number(first_value(form, "critical_power")),
        "training_effect_aerobic": parse_number(first_value(form, "training_effect_aerobic")),
        "training_effect_anaerobic": parse_number(first_value(form, "training_effect_anaerobic")),
        "training_load": parse_number(first_value(form, "training_load")),
        "recovery_time_hr": parse_number(first_value(form, "recovery_time_hr")),
        "notes": first_value(form, "notes"),
    }


def option_tags(options, selected=""):
    tags = ['<option value="">自動 / 留空</option>']
    for option in options:
        value = html.escape(str(option), quote=True)
        is_selected = " selected" if str(option) == selected else ""
        tags.append(f'<option value="{value}"{is_selected}>{html.escape(str(option))}</option>')
    return "\n".join(tags)


def multi_option_tags(options, selected=None):
    selected = set(selected or [])
    tags = []
    for option in options:
        value = html.escape(str(option), quote=True)
        is_selected = " selected" if str(option) in selected else ""
        tags.append(f'<option value="{value}"{is_selected}>{html.escape(str(option))}</option>')
    return "\n".join(tags)


def input_field(label, name, value="", input_type="text", placeholder=""):
    return f"""
      <label>
        <span>{html.escape(label)}</span>
        <input type="{input_type}" name="{html.escape(name)}" value="{html.escape(str(value or ''), quote=True)}" placeholder="{html.escape(placeholder, quote=True)}">
      </label>
    """


def workout_focus_reference_table(dropdown_options):
    mapping = dropdown_options.get(WORKOUT_FOCUS_MAP_KEY, {})
    rows = []
    for workout in dropdown_options["workout_types"]:
        focus_matches = mapping.get(workout, [])
        focus_label = "、".join(focus_matches) if focus_matches else "未設定"
        focus_data = "||".join(focus_matches)
        rows.append(
            f"""
            <tr>
              <td>{html.escape(workout)}</td>
              <td>{html.escape(focus_label)}</td>
              <td>
                <button class="small-button" type="button" data-workout-value="{html.escape(workout, quote=True)}" data-focus-values="{html.escape(focus_data, quote=True)}">套用</button>
              </td>
            </tr>
            """
        )
    return f"""
      <section class="reference">
        <div class="reference-head">
          <h2>課表與訓練目的對照</h2>
        </div>
        <div class="reference-table-wrap">
          <table class="reference-table">
            <thead>
              <tr>
                <th>課表類型</th>
                <th>預設訓練目的</th>
                <th></th>
              </tr>
            </thead>
            <tbody>
              {"".join(rows)}
            </tbody>
          </table>
        </div>
      </section>
    """


def product_banner(title="跑步分析資料轉檔"):
    return f"""
      <section class="product-banner">
        <div class="banner-copy">
          <div class="brand-row">
            <img class="brand-mark" src="/assets/rac_mark_transparent.png" alt="RAC">
            <div>
              <h1>{html.escape(title)}</h1>
              <p class="brand-subtitle">RUNNING ANALYTICS CONVERTER</p>
            </div>
          </div>
          <p class="banner-subtitle">Garmin FIT -> Standardized Excel -> AI Coach Analysis</p>
          <div class="flow">
            <span>FIT Import</span>
            <span>{html.escape(EXCEL_SCHEMA_LABEL)}</span>
            <span>AI Coach</span>
            <span>Long-term Analytics</span>
          </div>
          <p class="banner-note">將 Garmin FIT 活動檔轉換為固定格式 Excel，支援每日 AI 教練分析、週/月趨勢與長期跑步資料庫。</p>
        </div>
        <div class="version-panel">
          <span>App v{html.escape(APP_VERSION)}</span>
          <span>{html.escape(EXCEL_SCHEMA_LABEL)}</span>
        </div>
      </section>
    """


def nav(active="convert"):
    convert_class = " active" if active == "convert" else ""
    options_class = " active" if active == "options" else ""
    return f"""
      <nav>
        <a class="nav-link{convert_class}" href="/">轉檔</a>
        <a class="nav-link{options_class}" href="/options">下拉選單設定</a>
      </nav>
    """


def status_html(message="", error=""):
    if message:
        return f'<section class="status ok">{message}</section>'
    if error:
        return f'<section class="status error">{html.escape(error)}</section>'
    return ""


def base_styles():
    return """
    :root {
      color-scheme: light;
      --ink: #18222f;
      --muted: #657386;
      --line: #d9e3ee;
      --accent: #0f766e;
      --accent-dark: #0b4f5f;
      --accent-soft: #e7f4f2;
      --surface: #ffffff;
      --page: #f3f7fa;
      --error: #b42318;
      --ok: #166534;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Noto Sans TC", sans-serif;
      background:
        linear-gradient(180deg, #edf5f7 0, var(--page) 260px),
        var(--page);
      color: var(--ink);
    }
    main {
      width: min(1120px, calc(100vw - 32px));
      margin: 28px auto 40px;
    }
    h1 {
      margin: 0 0 6px;
      font-size: 30px;
      letter-spacing: 0;
    }
    .product-banner {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 24px;
      align-items: start;
      margin: 0 0 18px;
      padding: 28px 32px;
      border-radius: 18px;
      color: #fff;
      background:
        linear-gradient(90deg, rgba(3, 33, 48, 0.7) 0%, rgba(5, 55, 65, 0.38) 46%, rgba(5, 87, 88, 0.2) 100%),
        url("/assets/rac_banner.png") center / cover no-repeat;
      box-shadow: 0 18px 48px rgba(11, 79, 95, 0.22);
      min-height: 286px;
    }
    .eyebrow {
      margin: 0 0 8px;
      font-size: 13px;
      font-weight: 800;
      letter-spacing: 0.04em;
      text-transform: uppercase;
      color: rgba(255, 255, 255, 0.76);
    }
    .product-banner h1 {
      margin: 0;
      font-size: 42px;
      line-height: 1.15;
      text-shadow: 0 2px 12px rgba(0, 0, 0, 0.24);
    }
    .brand-row {
      display: flex;
      align-items: center;
      gap: 24px;
    }
    .brand-mark {
      width: min(320px, 32vw);
      height: auto;
      filter: drop-shadow(0 10px 20px rgba(0, 0, 0, 0.28));
    }
    .brand-subtitle {
      margin: 8px 0 0;
      color: rgba(220, 236, 242, 0.76);
      font-size: 20px;
      font-weight: 800;
      letter-spacing: 0.08em;
    }
    .banner-subtitle {
      margin: 18px 0 0;
      font-size: 16px;
      color: rgba(255, 255, 255, 0.86);
    }
    .banner-note {
      max-width: 720px;
      margin: 16px 0 0;
      color: rgba(255, 255, 255, 0.86);
      line-height: 1.6;
    }
    .subtitle {
      margin: 0 0 18px;
      color: var(--muted);
      font-size: 15px;
      line-height: 1.6;
    }
    .flow {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 18px 0 0;
    }
    .flow span {
      color: #fff;
      background: rgba(255, 255, 255, 0.14);
      border: 1px solid rgba(255, 255, 255, 0.2);
      border-radius: 999px;
      padding: 7px 10px;
      font-size: 13px;
      font-weight: 700;
    }
    .flow span + span::before {
      content: "-> ";
      opacity: 0.68;
    }
    .version {
      display: inline-block;
      margin-left: 8px;
      color: var(--muted);
      font-size: 14px;
      font-weight: 500;
    }
    .version-panel {
      display: grid;
      gap: 8px;
      min-width: 180px;
      justify-items: end;
      position: relative;
      z-index: 1;
    }
    .version-panel span {
      display: inline-flex;
      color: #fff;
      background: rgba(255, 255, 255, 0.16);
      border: 1px solid rgba(255, 255, 255, 0.24);
      border-radius: 999px;
      padding: 8px 11px;
      font-size: 13px;
      font-weight: 800;
      white-space: nowrap;
    }
    nav {
      display: flex;
      width: fit-content;
      gap: 4px;
      margin: 0 0 18px;
      padding: 4px;
      border: 1px solid var(--line);
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.74);
      box-shadow: 0 8px 24px rgba(31, 41, 51, 0.06);
    }
    .nav-link {
      color: var(--muted);
      text-decoration: none;
      padding: 9px 14px;
      border-radius: 999px;
      font-weight: 700;
    }
    .nav-link.active {
      color: #fff;
      background: var(--accent-dark);
    }
    form {
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 24px;
      box-shadow: 0 16px 42px rgba(31, 41, 51, 0.1);
    }
    fieldset {
      border: 0;
      padding: 0;
      margin: 0 0 24px;
    }
    legend {
      padding: 0;
      margin: 0 0 14px;
      font-size: 17px;
      font-weight: 700;
    }
    .grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 14px;
    }
    label {
      display: flex;
      flex-direction: column;
      gap: 6px;
      min-width: 0;
    }
    label.wide { grid-column: span 3; }
    span {
      font-size: 13px;
      color: var(--muted);
    }
    input, select, textarea {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px 11px;
      font: inherit;
      color: var(--ink);
      background: #fff;
    }
    input:focus, select:focus, textarea:focus {
      outline: 3px solid rgba(15, 118, 110, 0.16);
      border-color: var(--accent);
    }
    select[multiple] {
      min-height: 132px;
    }
    textarea {
      min-height: 76px;
      resize: vertical;
    }
    textarea.tall {
      min-height: 220px;
      line-height: 1.45;
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size: 13px;
    }
    .inline {
      display: flex;
      align-items: center;
      gap: 10px;
      color: var(--ink);
      margin-top: 2px;
    }
    .inline input {
      width: auto;
    }
    .actions {
      display: flex;
      align-items: center;
      gap: 12px;
      margin-top: 4px;
    }
    button, .button {
      appearance: none;
      border: 0;
      border-radius: 8px;
      background: var(--accent-dark);
      color: #fff;
      padding: 11px 16px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      text-decoration: none;
      display: inline-block;
    }
    button:hover, .button:hover {
      background: var(--accent);
    }
    .secondary {
      background: #e7f0f3;
      color: var(--ink);
    }
    .secondary:hover {
      background: #d7e3ee;
    }
    .status {
      border-radius: 8px;
      padding: 14px 16px;
      margin: 0 0 18px;
      border: 1px solid var(--line);
      background: #fff;
      line-height: 1.55;
    }
    .ok { color: var(--ok); }
    .error { color: var(--error); }
    .note {
      color: var(--muted);
      font-size: 13px;
      margin: 8px 0 0;
    }
    .summary {
      width: 100%;
      border-collapse: collapse;
      margin: 10px 0 14px;
      color: var(--ink);
    }
    .summary th,
    .summary td {
      border-bottom: 1px solid var(--line);
      padding: 8px 6px;
      text-align: left;
      vertical-align: top;
    }
    .summary th {
      width: 130px;
      color: var(--muted);
      font-weight: 700;
    }
    .reference {
      margin: 0 0 24px;
    }
    .reference-head {
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 12px;
      margin: 0 0 10px;
    }
    .reference h2 {
      margin: 0;
      font-size: 17px;
      letter-spacing: 0;
    }
    .reference-table-wrap {
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: #fff;
      box-shadow: 0 8px 22px rgba(31, 41, 51, 0.06);
    }
    .reference-table {
      width: 100%;
      min-width: 640px;
      border-collapse: collapse;
      color: var(--ink);
    }
    .reference-table th,
    .reference-table td {
      border-bottom: 1px solid var(--line);
      padding: 10px 12px;
      text-align: left;
      vertical-align: middle;
    }
    .reference-table th {
      background: #f7fafb;
      font-weight: 800;
    }
    .reference-table tr:last-child td {
      border-bottom: 0;
    }
    .small-button {
      padding: 7px 10px;
      font-size: 13px;
      white-space: nowrap;
    }
    code {
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size: 13px;
      color: var(--ink);
    }
    @media (max-width: 760px) {
      main { width: min(100vw - 20px, 1040px); margin: 18px auto; }
      .product-banner {
        grid-template-columns: 1fr;
        padding: 22px;
        border-radius: 14px;
        min-height: 0;
      }
      .brand-row {
        display: grid;
        gap: 12px;
      }
      .brand-mark {
        width: min(260px, 78vw);
      }
      .product-banner h1 { font-size: 30px; }
      .brand-subtitle { font-size: 14px; }
      .version-panel {
        justify-items: start;
        grid-template-columns: repeat(2, minmax(0, auto));
      }
      form { padding: 16px; }
      .grid { grid-template-columns: 1fr; }
      label.wide { grid-column: span 1; }
      .actions { flex-direction: column; align-items: stretch; }
      button, .button { text-align: center; }
      nav { overflow-x: auto; }
      .nav-link { white-space: nowrap; }
    }
    """


def render_page(message="", error="", selected_fit=""):
    dropdown_options = load_app_options()
    files, total_count, list_limit = fit_files(selected_fit)
    fit_options = []
    for path in files:
        value = html.escape(path.name, quote=True)
        is_selected = " selected" if path.name == selected_fit else ""
        fit_options.append(f'<option value="{value}"{is_selected}>{html.escape(path.name)}</option>')
    if not fit_options:
        fit_options.append('<option value="">FIT 資料夾目前沒有 .fit 檔</option>')

    list_note = f"FIT 資料夾共有 {total_count} 個檔案，目前清單只顯示最近 {min(total_count, list_limit)} 個。"

    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>跑步分析資料轉檔</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner()}
    {nav("convert")}
    {status_html(message, error)}
    <form method="post" action="/convert" enctype="multipart/form-data">
      <fieldset>
        <legend>檔案</legend>
        <div class="grid">
          <label class="wide">
            <span>從電腦選擇 FIT 檔</span>
            <input type="file" name="upload_fit" accept=".fit">
          </label>
          <label class="wide">
            <span>或使用 FIT 資料夾裡的檔案</span>
            <select name="fit_file">
              {"".join(fit_options)}
            </select>
            <p class="note">{html.escape(list_note)}</p>
          </label>
          <label class="wide">
            <span>輸出檔名，可留空</span>
            <input name="output_name" placeholder="{html.escape(WORKBOOK_VERSION_NAME)}_活動檔名.xlsx">
          </label>
          <label class="inline wide">
            <input type="checkbox" name="fetch_weather" value="1" checked>
            <span>自動抓 Open-Meteo 歷史天氣</span>
          </label>
        </div>
      </fieldset>

      <fieldset>
        <legend>活動資訊</legend>
        <div class="grid">
          <label>
            <span>鞋款</span>
            <select name="shoe">{option_tags(dropdown_options["shoes"])}</select>
          </label>
          <label>
            <span>課表類型</span>
            <select name="workout_type">{option_tags(dropdown_options["workout_types"])}</select>
          </label>
          <label>
            <span>訓練目的</span>
            <select name="training_focus" multiple>{multi_option_tags(dropdown_options["training_focus"])}</select>
            <p class="note">可多選；macOS 按 Command，Windows 按 Ctrl。</p>
          </label>
          <label>
            <span>Garmin 主觀感受</span>
            <select name="rpe">{option_tags(dropdown_options["garmin_rpe"])}</select>
          </label>
          {input_field("最大心率", "max_hr", input_type="number", placeholder="自動")}
          {input_field("Critical Power(W)", "critical_power", input_type="number", placeholder="自動")}
          {input_field("氣溫(°C)", "weather_temp", input_type="number", placeholder="自動")}
          {input_field("濕度(%)", "humidity", input_type="number", placeholder="自動")}
          {input_field("風向", "wind_direction", placeholder="自動")}
          {input_field("風速", "wind_speed", placeholder="自動")}
          {input_field("Recovery Time (hr)", "recovery_time_hr", input_type="number")}
          {input_field("Training Load", "training_load", input_type="number", placeholder="自動")}
          <label class="wide">
            <span>補給紀錄</span>
            <textarea name="fueling"></textarea>
          </label>
          <label class="wide">
            <span>備註</span>
            <textarea name="notes"></textarea>
          </label>
        </div>
      </fieldset>

      {workout_focus_reference_table(dropdown_options)}

      <div class="actions">
        <button type="submit">轉成 Excel</button>
        <a class="button secondary" href="/">重新整理</a>
      </div>
    </form>
  </main>
  <script>
    function chooseOption(select, optionValue, allowMultiple) {{
      if (!select || !optionValue) return;
      for (const option of select.options) {{
        if (option.value === optionValue) {{
          if (allowMultiple) {{
            option.selected = true;
          }} else {{
            select.value = option.value;
          }}
          return;
        }}
      }}
    }}

    document.querySelectorAll("[data-workout-value]").forEach((button) => {{
      button.addEventListener("click", () => {{
        const workoutSelect = document.querySelector('select[name="workout_type"]');
        const focusSelect = document.querySelector('select[name="training_focus"]');
        chooseOption(workoutSelect, button.dataset.workoutValue, false);
        if (focusSelect) {{
          for (const option of focusSelect.options) option.selected = false;
          button.dataset.focusValues.split("||").forEach((focus) => chooseOption(focusSelect, focus, true));
        }}
      }});
    }});
  </script>
</body>
</html>"""


def options_textarea(name, label, values):
    text = "\n".join(str(value) for value in values)
    return f"""
      <label class="wide">
        <span>{html.escape(label)}，每行一個選項</span>
        <textarea class="tall" name="{html.escape(name)}">{html.escape(text)}</textarea>
      </label>
    """


def mapping_select(name, training_focus, selected):
    return f"""
      <select name="{html.escape(name)}" multiple>
        {multi_option_tags(training_focus, selected)}
      </select>
    """


def mapping_settings_table(options):
    rows = []
    mapping = options.get(WORKOUT_FOCUS_MAP_KEY, {})
    training_focus = options["training_focus"]
    for index, workout in enumerate(options["workout_types"]):
        selected = mapping.get(workout, [])
        rows.append(
            f"""
            <tr>
              <td>
                {html.escape(workout)}
                <input type="hidden" name="map_workout_{index}" value="{html.escape(workout, quote=True)}">
              </td>
              <td>{mapping_select(f"map_focus_{index}", training_focus, selected)}</td>
            </tr>
            """
        )
    return f"""
      <div class="reference-table-wrap">
        <table class="reference-table">
          <thead>
            <tr>
              <th>課表類型</th>
              <th>對應訓練目的</th>
            </tr>
          </thead>
          <tbody>
            {"".join(rows)}
          </tbody>
        </table>
      </div>
    """


def dropdown_options_from_form(form):
    result = {}
    for key, _label in OPTION_FIELDS:
        lines = [line.strip() for line in first_value(form, key).splitlines()]
        values = []
        seen = set()
        for line in lines:
            if not line or line in seen:
                continue
            values.append(line)
            seen.add(line)
        if not values:
            raise ValueError("每一組下拉選單至少需要一個選項。")
        result[key] = values
    result[WORKOUT_FOCUS_MAP_KEY] = workout_focus_map_from_form(form, result)
    return result


def workout_focus_map_from_form(form, options):
    valid_workouts = set(options["workout_types"])
    valid_focus = set(options["training_focus"])
    result = {}
    for index, workout in enumerate(options["workout_types"]):
        posted_workout = first_value(form, f"map_workout_{index}") or workout
        if posted_workout not in valid_workouts:
            continue
        selected = [
            value
            for value in selected_values(form, f"map_focus_{index}")
            if value in valid_focus
        ]
        result[posted_workout] = selected
    defaults = default_workout_focus_map(options)
    for workout in options["workout_types"]:
        result.setdefault(workout, defaults.get(workout, []))
    return result


def save_dropdown_options(options):
    CONFIG_DIR = DROPDOWN_CONFIG_PATH.parent
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    DROPDOWN_CONFIG_PATH.write_text(
        json.dumps(options, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def render_options_page(message="", error=""):
    options = load_app_options()
    fields = "\n".join(options_textarea(key, label, options[key]) for key, label in OPTION_FIELDS)
    return f"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>下拉選單設定</title>
  <style>
    {base_styles()}
  </style>
</head>
<body>
  <main>
    {product_banner("下拉選單設定")}
    <p class="subtitle">先修改活動資訊選項並儲存，再設定課表類型與訓練目的的對應關係。儲存後會立即套用到轉檔頁與輸出的 Excel。</p>
    {nav("options")}
    {status_html(message, error)}
    <form method="post" action="/options">
      <fieldset>
        <legend>選項內容</legend>
        <div class="grid">
          {fields}
        </div>
      </fieldset>
      <fieldset>
        <legend>課表與訓練目的對應</legend>
        <p class="note">若剛新增或改名課表/訓練目的，請先儲存選項內容；頁面重新整理後再設定對應關係。</p>
        {mapping_settings_table(options)}
      </fieldset>
      <div class="actions">
        <button type="submit">儲存設定</button>
        <a class="button secondary" href="/">回轉檔</a>
      </div>
    </form>
  </main>
</body>
</html>"""


class AppHandler(BaseHTTPRequestHandler):
    def send_html(self, content, status=200):
        data = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_xlsx(self, path):
        data = path.read_bytes()
        filename = path.name
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"attachment; filename*=UTF-8''{quote(filename)}",
        )
        self.end_headers()
        self.wfile.write(data)

    def send_asset(self, path):
        try:
            resolved = path.resolve()
            asset_root = ASSETS_DIR.resolve()
            if not resolved.is_file() or not resolved.is_relative_to(asset_root):
                raise FileNotFoundError
            data = resolved.read_bytes()
        except (OSError, FileNotFoundError):
            self.send_html(render_page(error="找不到指定的介面資產。"), status=404)
            return
        content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path.startswith("/assets/"):
            self.send_asset(ASSETS_DIR / parsed.path.removeprefix("/assets/"))
            return
        if parsed.path == "/options":
            self.send_html(render_options_page())
            return
        if parsed.path == "/open":
            output = Path(first_value(query, "path"))
            if is_output_file(output):
                open_file(output)
                self.send_html(render_page(message=f"已要求系統開啟 <code>{html.escape(str(output))}</code>"))
            else:
                self.send_html(render_page(error="找不到輸出檔。"), status=404)
            return
        if parsed.path == "/download":
            output = Path(first_value(query, "path"))
            if is_output_file(output):
                self.send_xlsx(output)
            else:
                self.send_html(render_page(error="找不到可下載的 Excel 檔。"), status=404)
            return
        if parsed.path == "/open-folder":
            DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            open_file(DEFAULT_OUTPUT_DIR)
            self.send_html(render_page(message=f"已要求系統開啟 <code>{html.escape(str(DEFAULT_OUTPUT_DIR))}</code>"))
            return
        self.send_html(render_page())

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/options":
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            form, _files = parse_post_data(self.headers, body)
            try:
                options = dropdown_options_from_form(form)
                save_dropdown_options(options)
            except Exception as error:
                self.send_html(render_options_page(error=f"儲存失敗：{error}"), status=400)
                return
            self.send_html(render_options_page(message="下拉選單已更新。"))
            return

        if parsed.path != "/convert":
            self.send_html(render_page(error="不支援的操作。"), status=404)
            return

        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        try:
            form, files = parse_post_data(self.headers, body)
            fit_path = save_uploaded_fit(files.get("upload_fit"))
        except ValueError as error:
            self.send_html(render_page(error=str(error)), status=400)
            return

        fit_name = first_value(form, "fit_file")
        if fit_path is None:
            fit_path = FIT_DIR / fit_name
        if not fit_path.exists() or fit_path.suffix.lower() != ".fit":
            self.send_html(render_page(error="請選擇一個 .fit 檔，或從 FIT 資料夾清單選擇有效檔案。"), status=400)
            return
        fit_name = fit_path.name

        output_name = first_value(form, "output_name")
        output_name = Path(output_name).name if output_name else ""
        output_path = DEFAULT_OUTPUT_DIR / output_name if output_name else DEFAULT_OUTPUT_DIR / f"{WORKBOOK_VERSION_NAME}_{fit_path.stem}.xlsx"
        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")

        metadata = build_metadata(form)
        fetch_weather = first_value(form, "fetch_weather") == "1"

        try:
            saved = create_workbook(fit_path, output_path, metadata=metadata, fetch_weather=fetch_weather)
        except Exception as error:
            self.send_html(render_page(error=friendly_error(error), selected_fit=fit_name), status=500)
            return

        open_link = "/open?" + urlencode({"path": str(saved)})
        download_link = "/download?" + urlencode({"path": str(saved)})
        folder_link = "/open-folder"
        try:
            summary = summary_html(workbook_summary(saved))
        except Exception:
            summary = ""
        message = (
            f"轉檔完成：<code>{html.escape(str(saved))}</code><br>"
            f"{summary}"
            f'<a class="button" href="{html.escape(open_link, quote=True)}">開啟 Excel</a> '
            f'<a class="button secondary" href="{html.escape(download_link, quote=True)}">下載 Excel</a> '
            f'<a class="button secondary" href="{html.escape(folder_link, quote=True)}">開啟 EXCEL 資料夾</a>'
        )
        self.send_html(render_page(message=message, selected_fit=fit_name))

    def log_message(self, format, *args):
        return


def open_browser_later(url):
    timer = threading.Timer(0.6, lambda: webbrowser.open(url))
    timer.daemon = True
    timer.start()


def main():
    try:
        server = ThreadingHTTPServer((HOST, PORT), AppHandler)
    except PermissionError:
        print(f"無法啟動本機網站：系統目前不允許使用 {HOST}:{PORT}。")
        print("請確認防火牆或安全性設定，或改用 CLI 方式轉檔。")
        return
    except OSError as error:
        if getattr(error, "errno", None) == 48:
            print(f"無法啟動本機網站：{HOST}:{PORT} 已經被其他程式使用。")
            print("請關掉舊的轉檔視窗，或稍後再重新啟動。")
        else:
            print(f"無法啟動本機網站：{error}")
        return
    url = f"http://{HOST}:{PORT}"
    print(f"Running Analytics v{APP_VERSION}: {url}")
    open_browser_later(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__ == "__main__":
    main()
