#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from statistics import mean
from urllib.parse import urlencode
from urllib.request import urlopen

from garmin_fit_sdk import Decoder, Stream
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FIT_EPOCH = 631065600
APP_VERSION = "1.5.0"
WORKBOOK_VERSION_NAME = "跑步分析資料 v1.1"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "EXCEL"
CONFIG_DIR = Path(__file__).resolve().parent / "config"
DROPDOWN_CONFIG_PATH = CONFIG_DIR / "dropdown_options.json"
STAMINA_RECORD_FIELDS = (137, 138)
STAMINA_SESSION_START = 205
STAMINA_SESSION_END_FIELDS = (206, 207)

DEFAULT_DROPDOWN_OPTIONS = {
    "shoes": [
    "Boston 13 Green",
    "Boston 13 Blue",
    "EVO SL",
    "Rebel v5",
    "Nimbus 28",
    ],
    "workout_types": [
    "Recovery Run（恢復跑）",
    "Easy Run（輕鬆跑）",
    "LSD（長距離慢跑）",
    "Long Run（長跑）",
    "Tempo Run（節奏跑）",
    "Marathon Pace（馬拉松配速）",
    "Interval（間歇）",
    "Repetition（速度訓練）",
    "Progression Run（漸速跑）",
    "Fartlek（法特萊克）",
    "Race（比賽）",
    "Other（其他）",
    ],
    "training_focus": [
    "Recovery",
    "Aerobic Base",
    "Endurance",
    "Marathon Pace",
    "Threshold",
    "VO₂max",
    "Speed",
    "Running Economy",
    "Heat Adaptation",
    "Race",
    ],
    "garmin_rpe": [
    "1 - 非常輕鬆",
    "2 - 輕鬆",
    "3 - 中等",
    "4 - 有點難",
    "5 - 困難",
    "6 - 困難",
    "7 - 非常困難",
    "8 - 非常困難",
    "9 - 超級難",
    "10 - 極限",
    ],
    "garmin_feel": [
    "非常弱",
    "弱",
    "普通",
    "強",
    "非常強",
    ],
}
WEATHER_FIELDS = ("weather_temp", "humidity", "wind_direction", "wind_speed", "weather_description")


HEADERS = [
    "公里",
    "距離(m)",
    "時間(秒)",
    "配速(分:秒/km)",
    "平均心率",
    "平均心率%",
    "最高心率",
    "平均步頻(spm)",
    "平均功率(W)",
    "平均功率%",
    "垂直振幅(mm)",
    "垂直比(%)",
    "觸地時間(ms)",
    "步幅(mm)",
    "溫度(°C)",
    "Stamina 起",
    "Stamina 末",
    "爬升(m)",
]


def fit_datetime(value):
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, (int, float)):
        return dt.datetime.fromtimestamp(value + FIT_EPOCH, tz=dt.timezone.utc)
    return None


def pace_text(seconds: float, meters: float) -> str:
    if not meters:
        return ""
    sec_per_km = round(seconds / meters * 1000)
    return f"{sec_per_km // 60}:{sec_per_km % 60:02d}"


def rounded(value, ndigits=1):
    if value is None:
        return None
    return round(float(value), ndigits)


def cell_value(value):
    return "" if value is None else value


def average(values, ndigits=1):
    vals = [float(v) for v in values if isinstance(v, (int, float))]
    return round(mean(vals), ndigits) if vals else None


def maximum(values):
    vals = [float(v) for v in values if isinstance(v, (int, float))]
    if not vals:
        return None
    value = max(vals)
    return int(value) if value.is_integer() else value


def first_number(row, *fields):
    for field in fields:
        value = row.get(field)
        if isinstance(value, (int, float)):
            return value
    return None


def parse_garmin_activity_id(path: Path):
    numbers = re.findall(r"\d{10,}", path.stem)
    return int(numbers[0]) if numbers else ""


def fit_sha256(path: Path):
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def first_session(messages):
    sessions = messages.get("session_mesgs", []) if messages else []
    return sessions[0] if sessions else {}


def path_date_code(path: Path):
    for match in re.finditer(r"(\d{4})(\d{2})(\d{2})", path.stem):
        month = int(match.group(2))
        day = int(match.group(3))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return "".join(match.groups())
    return ""


def fit_activity_date_code(path: Path, messages=None):
    if messages is None:
        existing = path_date_code(path)
        if existing:
            return existing
        messages = decode_fit(path)
    session = first_session(messages)
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start:
        return start.astimezone().strftime("%Y%m%d")
    return path_date_code(path) or dt.date.today().strftime("%Y%m%d")


def fit_activity_identifier(path: Path):
    activity_id = parse_garmin_activity_id(path)
    if activity_id:
        return str(activity_id)
    return f"FIT{fit_sha256(path)[:12]}"


def output_file_stem(path: Path, messages=None):
    if messages is None and path_date_code(path) and parse_garmin_activity_id(path):
        return f"{path_date_code(path)}_{fit_activity_identifier(path)}"
    return f"{fit_activity_date_code(path, messages)}_{fit_activity_identifier(path)}"


def output_month_label(path: Path, messages=None):
    if re.fullmatch(r"\d{4}-\d{2}", path.parent.name):
        return path.parent.name
    existing = path_date_code(path)
    if existing:
        return f"{existing[:4]}-{existing[4:6]}"
    date_code = fit_activity_date_code(path, messages)
    return f"{date_code[:4]}-{date_code[4:6]}"


def default_output_path(fit_path: Path):
    messages = None if path_date_code(fit_path) and parse_garmin_activity_id(fit_path) else decode_fit(fit_path)
    return DEFAULT_OUTPUT_DIR / output_month_label(fit_path, messages) / f"{WORKBOOK_VERSION_NAME}_{output_file_stem(fit_path, messages)}.xlsx"


def load_dropdown_options(path=DROPDOWN_CONFIG_PATH):
    options = {key: list(value) for key, value in DEFAULT_DROPDOWN_OPTIONS.items()}
    if not path.exists():
        return options
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        print(f"Dropdown config ignored: {error}")
        return options

    for key, default_values in DEFAULT_DROPDOWN_OPTIONS.items():
        values = loaded.get(key)
        if isinstance(values, list):
            cleaned = [str(value).strip() for value in values if str(value).strip()]
            if cleaned:
                options[key] = cleaned
        if not options.get(key):
            options[key] = list(default_values)
    return options


def semicircles_to_degrees(value):
    if not isinstance(value, (int, float)):
        return None
    return float(value) * 180.0 / 2**31


def decode_fit(path: Path):
    stream = Stream.from_file(str(path))
    messages, errors = Decoder(stream).read()
    if errors:
        raise RuntimeError(f"FIT decode errors: {errors}")
    return messages


def compass_direction(degrees):
    if not isinstance(degrees, (int, float)):
        return ""
    labels = [
        "北風",
        "東北偏北風",
        "東北風",
        "東北偏東風",
        "東風",
        "東南偏東風",
        "東南風",
        "東南偏南風",
        "南風",
        "西南偏南風",
        "西南風",
        "西南偏西風",
        "西風",
        "西北偏西風",
        "西北風",
        "西北偏北風",
    ]
    index = int((float(degrees) + 11.25) // 22.5) % 16
    return labels[index]


def weather_code_description(code):
    descriptions = {
        0: "晴",
        1: "大致晴朗",
        2: "局部多雲",
        3: "陰",
        45: "霧",
        48: "霧淞",
        51: "毛毛雨",
        53: "毛毛雨",
        55: "毛毛雨",
        56: "凍毛毛雨",
        57: "凍毛毛雨",
        61: "小雨",
        63: "中雨",
        65: "大雨",
        66: "凍雨",
        67: "凍雨",
        71: "小雪",
        73: "中雪",
        75: "大雪",
        77: "雪粒",
        80: "陣雨",
        81: "陣雨",
        82: "強陣雨",
        85: "陣雪",
        86: "強陣雪",
        95: "雷雨",
        96: "雷雨伴冰雹",
        99: "雷雨伴冰雹",
    }
    if not isinstance(code, (int, float)):
        return ""
    return descriptions.get(int(code), "")


def nearest_hour_index(times, target):
    best_index = None
    best_seconds = None
    for index, value in enumerate(times):
        try:
            current = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            continue
        if current.tzinfo is None:
            current = current.replace(tzinfo=dt.timezone.utc)
        delta = abs((current.astimezone(dt.timezone.utc) - target).total_seconds())
        if best_seconds is None or delta < best_seconds:
            best_index = index
            best_seconds = delta
    return best_index


def activity_location(session, records):
    lat = semicircles_to_degrees(session.get("start_position_lat") or session.get("end_position_lat"))
    lon = semicircles_to_degrees(session.get("start_position_long") or session.get("end_position_long"))
    if lat is not None and lon is not None:
        return lat, lon

    for record in records:
        lat = semicircles_to_degrees(record.get("position_lat"))
        lon = semicircles_to_degrees(record.get("position_long"))
        if lat is not None and lon is not None:
            return lat, lon
    return None, None


def fetch_weather_for_activity(session, records):
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start is None:
        return {}
    start_utc = start.astimezone(dt.timezone.utc)
    latitude, longitude = activity_location(session, records)
    if latitude is None or longitude is None:
        return {}

    query = {
        "latitude": round(latitude, 6),
        "longitude": round(longitude, 6),
        "start_date": start_utc.date().isoformat(),
        "end_date": start_utc.date().isoformat(),
        "hourly": ",".join(
            [
                "temperature_2m",
                "relative_humidity_2m",
                "wind_speed_10m",
                "wind_direction_10m",
                "weather_code",
            ]
        ),
        "timezone": "UTC",
        "wind_speed_unit": "kmh",
    }
    url = "https://archive-api.open-meteo.com/v1/archive?" + urlencode(query)
    with urlopen(url, timeout=15) as response:
        payload = json.load(response)

    hourly = payload.get("hourly") or {}
    index = nearest_hour_index(hourly.get("time") or [], start_utc)
    if index is None:
        return {}

    def hourly_value(name):
        values = hourly.get(name) or []
        if index >= len(values):
            return None
        return values[index]

    wind_degrees = hourly_value("wind_direction_10m")
    wind_label = compass_direction(wind_degrees)
    weather = {
        "weather_temp": rounded(hourly_value("temperature_2m"), 1),
        "humidity": rounded(hourly_value("relative_humidity_2m"), 0),
        "wind_direction": f"{round(wind_degrees)}° ({wind_label})" if wind_degrees is not None else "",
        "wind_speed": f"{rounded(hourly_value('wind_speed_10m'), 1)} km/h"
        if hourly_value("wind_speed_10m") is not None
        else "",
        "weather_description": weather_code_description(hourly_value("weather_code")),
    }
    return {key: value for key, value in weather.items() if value not in ("", None)}


def records_for_lap(records, start_time, elapsed_seconds):
    if not start_time or not elapsed_seconds:
        return []
    end_time = start_time + dt.timedelta(seconds=float(elapsed_seconds))
    return [
        record
        for record in records
        if (timestamp := fit_datetime(record.get("timestamp")))
        and start_time <= timestamp <= end_time
    ]


def stamina_at(records, fallback=None):
    for record in records:
        value = first_number(record, *STAMINA_RECORD_FIELDS)
        if value is not None:
            return int(value)
    return fallback


def first_message_number(messages, message_names, *fields):
    for message_name in message_names:
        for message in messages.get(message_name, []):
            value = first_number(message, *fields)
            if value is not None:
                return value
    return None


def garmin_rpe_label(value, rpe_options):
    if not isinstance(value, (int, float)):
        return ""
    rating = int(round(float(value) / 10))
    if 1 <= rating <= len(rpe_options):
        return rpe_options[rating - 1]
    return str(value)


def normalize_rpe(value, rpe_options):
    if value in ("", None):
        return ""
    if isinstance(value, str):
        if value in rpe_options:
            return value
        for option in rpe_options:
            label = option.split(" - ", 1)[-1]
            if value == label or value == label.split(" (", 1)[0]:
                return option
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)

    rating = int(round(numeric / 10)) if numeric > len(rpe_options) else int(round(numeric))
    if 1 <= rating <= len(rpe_options):
        return rpe_options[rating - 1]
    return str(value)


def garmin_feel_label(value, feel_options):
    if not isinstance(value, (int, float)):
        return ""
    scores = [0, 25, 50, 75, 100]
    index = min(range(len(scores)), key=lambda idx: abs(float(value) - scores[idx]))
    if index < len(feel_options):
        return feel_options[index]
    return str(value)


def normalize_feel(value, feel_options):
    if value in ("", None):
        return ""
    if isinstance(value, str):
        if value in feel_options:
            return value
        for option in feel_options:
            if value == option.split(" (", 1)[0]:
                return option
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    return garmin_feel_label(numeric, feel_options)


def build_rows(messages):
    laps = messages.get("lap_mesgs", [])
    records = messages.get("record_mesgs", [])
    sessions = messages.get("session_mesgs", [])
    session = sessions[0] if sessions else {}

    rows = []
    for index, lap in enumerate(laps, start=1):
        distance = float(lap.get("total_distance") or 0)
        elapsed = float(lap.get("total_timer_time") or lap.get("total_elapsed_time") or 0)
        lap_start = fit_datetime(lap.get("start_time"))
        lap_records = records_for_lap(records, lap_start, elapsed)

        cadence = first_number(lap, "avg_running_cadence", "avg_cadence")
        fractional_cadence = first_number(lap, "avg_fractional_cadence") or 0
        cadence_spm = (float(cadence) + float(fractional_cadence)) * 2 if cadence is not None else None

        start_stamina = stamina_at(lap_records)
        end_stamina = stamina_at(reversed(lap_records))
        if index == 1:
            start_stamina = int(first_number(session, STAMINA_SESSION_START) or start_stamina or 0)
        if index == len(laps):
            end_stamina = int(first_number(session, *STAMINA_SESSION_END_FIELDS) or end_stamina or 0)

        avg_heart_rate = first_number(lap, "avg_heart_rate")
        if avg_heart_rate is None:
            avg_heart_rate = average([record.get("heart_rate") for record in lap_records], 1)
        max_heart_rate = first_number(lap, "max_heart_rate")
        if max_heart_rate is None:
            max_heart_rate = maximum([record.get("heart_rate") for record in lap_records])
        avg_power = first_number(lap, "avg_power")
        if avg_power is None:
            avg_power = average([record.get("power") for record in lap_records], 1)

        rows.append(
            [
                index,
                round(distance),
                round(elapsed),
                pace_text(elapsed, distance),
                rounded(avg_heart_rate, 1),
                None,
                max_heart_rate,
                rounded(cadence_spm, 1),
                rounded(avg_power, 1),
                None,
                rounded(lap.get("avg_vertical_oscillation"), 1),
                rounded(lap.get("avg_vertical_ratio"), 1),
                rounded(lap.get("avg_stance_time"), 1),
                rounded(lap.get("avg_step_length"), 1),
                rounded(average([r.get("temperature") for r in lap_records], 1), 1)
                if lap_records
                else rounded(lap.get("avg_temperature"), 1),
                start_stamina,
                end_stamina,
                rounded(lap.get("total_ascent") or 0, 1),
            ]
        )
    return rows, session


def prompt_choice(label, options):
    print(f"\n{label}")
    for index, option in enumerate(options, start=1):
        print(f"  {index}. {option}")
    raw = input("請輸入編號或直接輸入文字，留空略過: ").strip()
    if not raw:
        return ""
    if raw.isdigit() and 1 <= int(raw) <= len(options):
        return options[int(raw) - 1]
    return raw


def prompt_text(label):
    return input(f"{label}，留空略過: ").strip()


def collect_metadata(args, dropdown_options):
    metadata = {
        "activity_name": args.activity_name or "",
        "shoe": args.shoe or "",
        "weather_temp": args.weather_temp if args.weather_temp is not None else "",
        "humidity": args.humidity if args.humidity is not None else "",
        "wind_direction": args.wind_direction or "",
        "wind_speed": args.wind_speed or "",
        "weather_description": args.weather_description or "",
        "workout_type": args.workout_type or "",
        "training_focus": args.training_focus or "",
        "feel": normalize_feel(args.feel, dropdown_options["garmin_feel"]),
        "rpe": normalize_rpe(args.rpe, dropdown_options["garmin_rpe"]),
        "fueling": args.fueling or "",
        "max_hr": args.max_hr if args.max_hr is not None else "",
        "critical_power": args.critical_power if args.critical_power is not None else "",
        "training_effect_aerobic": args.training_effect_aerobic if args.training_effect_aerobic is not None else "",
        "training_effect_anaerobic": args.training_effect_anaerobic if args.training_effect_anaerobic is not None else "",
        "training_load": args.training_load if args.training_load is not None else "",
        "recovery_time_hr": args.recovery_time_hr if args.recovery_time_hr is not None else "",
        "notes": args.notes or "",
    }
    if not args.interactive:
        return metadata

    if not metadata["activity_name"]:
        metadata["activity_name"] = prompt_text("活動名稱")
    if not metadata["shoe"]:
        metadata["shoe"] = prompt_choice("鞋款", dropdown_options["shoes"])
    if not args.fetch_weather and metadata["weather_temp"] == "":
        metadata["weather_temp"] = prompt_text("天氣氣溫(°C)")
    if not args.fetch_weather and metadata["humidity"] == "":
        metadata["humidity"] = prompt_text("濕度(%)")
    if not args.fetch_weather and not metadata["wind_direction"]:
        metadata["wind_direction"] = prompt_text("風向")
    if not args.fetch_weather and not metadata["wind_speed"]:
        metadata["wind_speed"] = prompt_text("風速")
    if not args.fetch_weather and not metadata["weather_description"]:
        metadata["weather_description"] = prompt_text("天氣描述")
    if not metadata["workout_type"]:
        metadata["workout_type"] = prompt_choice("課表類型", dropdown_options["workout_types"])
    if not metadata["training_focus"]:
        metadata["training_focus"] = prompt_choice("訓練目的（Training Focus）", dropdown_options["training_focus"])
    if metadata["feel"] == "":
        metadata["feel"] = prompt_text("感覺如何")
    if metadata["rpe"] == "":
        metadata["rpe"] = prompt_text("感受難度(1-10)")
    if not metadata["fueling"]:
        metadata["fueling"] = prompt_text("補給紀錄")
    if metadata["max_hr"] == "":
        metadata["max_hr"] = prompt_text("最大心率")
    if metadata["critical_power"] == "":
        metadata["critical_power"] = prompt_text("Critical Power(W)")
    if metadata["training_effect_aerobic"] == "":
        metadata["training_effect_aerobic"] = prompt_text("Training Effect (Aerobic)")
    if metadata["training_effect_anaerobic"] == "":
        metadata["training_effect_anaerobic"] = prompt_text("Training Effect (Anaerobic)")
    if metadata["training_load"] == "":
        metadata["training_load"] = prompt_text("Training Load")
    if metadata["recovery_time_hr"] == "":
        metadata["recovery_time_hr"] = prompt_text("Recovery Time (hr)")
    if not metadata["notes"]:
        metadata["notes"] = prompt_text("備註")
    return metadata


def apply_auto_weather(metadata, session, records, enabled):
    if not enabled:
        return metadata
    if all(metadata.get(key) not in ("", None) for key in WEATHER_FIELDS):
        return metadata
    try:
        weather = fetch_weather_for_activity(session, records)
    except Exception as error:
        print(f"Weather lookup skipped: {error}")
        return metadata
    result = dict(metadata)
    for key, value in weather.items():
        if result.get(key) in ("", None):
            result[key] = value
    return result


def apply_fit_metadata(metadata, messages):
    result = dict(metadata)
    if result.get("max_hr") in ("", None):
        value = first_message_number(
            messages,
            ("zones_target_mesgs", "time_in_zone_mesgs"),
            "max_heart_rate",
        )
        if value is not None:
            result["max_hr"] = value
    if result.get("critical_power") in ("", None):
        value = first_message_number(
            messages,
            ("zones_target_mesgs", "time_in_zone_mesgs"),
            "functional_threshold_power",
        )
        if value is not None:
            result["critical_power"] = value
    return result


def apply_file_identity(metadata, fit_path: Path):
    result = dict(metadata)
    result["garmin_activity_id"] = parse_garmin_activity_id(fit_path)
    result["fit_sha256"] = fit_sha256(fit_path)
    return result


def coerce_metadata(metadata):
    result = dict(metadata)
    for key in (
        "weather_temp",
        "humidity",
        "rpe",
        "max_hr",
        "critical_power",
        "training_effect_aerobic",
        "training_effect_anaerobic",
        "training_load",
        "recovery_time_hr",
    ):
        if result.get(key) == "":
            continue
        try:
            value = float(result.get(key, ""))
            result[key] = int(value) if value.is_integer() else value
        except (TypeError, ValueError):
            pass
    return result


def activity_date(session, fit_path: Path):
    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    if start:
        local = start.astimezone()
        return local.strftime("%Y/%m/%d")
    return fit_path.stem


def activity_type(session):
    profile_name = session.get("sport_profile_name")
    if profile_name not in ("", None):
        return str(profile_name)

    sport = session.get("sport")
    sub_sport = session.get("sub_sport")
    sport_labels = {
        "running": "跑步",
        "cycling": "自行車",
        "walking": "健走",
        "hiking": "健行",
        "swimming": "游泳",
    }
    sub_sport_labels = {
        "trail": "越野跑",
        "trail_running": "越野跑",
        "treadmill": "跑步機",
        "track": "田徑場跑步",
    }
    if sub_sport in sub_sport_labels:
        return sub_sport_labels[sub_sport]
    if sport in sport_labels:
        return sport_labels[sport]
    values = [str(value) for value in (sport, sub_sport) if value not in ("", None, "generic")]
    return " / ".join(values)


def activity_name(session, metadata):
    if metadata.get("activity_name") not in ("", None):
        return metadata.get("activity_name")
    for key in ("activity_name", "name", "workout_name", "title"):
        value = session.get(key)
        if value not in ("", None):
            return value
    return ""


def duration_text(seconds):
    if not isinstance(seconds, (int, float)):
        return ""
    total = int(round(float(seconds)))
    hours, remainder = divmod(total, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def activity_summary(rows):
    total_distance = sum(row[1] for row in rows if isinstance(row[1], (int, float)))
    total_seconds = sum(row[2] for row in rows if isinstance(row[2], (int, float)))
    return {
        "distance_km": round(total_distance / 1000, 2) if total_distance else "",
        "duration": duration_text(total_seconds),
        "avg_pace": pace_text(total_seconds, total_distance) if total_distance and total_seconds else "",
    }


def running_economy_summary(rows):
    return {
        "avg_cadence": average([row[7] for row in rows], 1),
        "avg_step_length": average([row[13] for row in rows], 1),
        "avg_gct": average([row[12] for row in rows], 1),
        "avg_vertical_oscillation": average([row[10] for row in rows], 1),
        "avg_vertical_ratio": average([row[11] for row in rows], 1),
    }


def stamina_summary(rows):
    if not rows:
        return "", ""
    return rows[0][15], rows[-1][16]


def apply_styles(ws, last_row):
    blue = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9E2F3")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws["A1"]
    title.fill = blue
    title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for cell in ws[2]:
        cell.fill = blue
        cell.font = Font(name="Noto Sans CJK SC", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin_gray, bottom=thin_gray)

    for row in ws.iter_rows(min_row=3, max_row=last_row, max_col=len(HEADERS)):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)

    for cell in ws[last_row]:
        cell.fill = total_fill
        cell.font = Font(name="Arial", size=10, bold=True)

    widths = {
        "A": 7,
        "B": 9,
        "C": 9,
        "D": 13,
        "E": 9,
        "F": 10,
        "G": 9,
        "H": 12,
        "I": 10,
        "J": 9,
        "K": 11,
        "L": 9,
        "M": 11,
        "N": 10,
        "O": 9,
        "P": 10,
        "Q": 10,
        "R": 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col in range(2, len(HEADERS) + 1):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0.0"
    for col in (1, 2, 3, 7, 16, 17):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0"
    for col in (6, 10):
        for row in range(3, last_row + 1):
            ws.cell(row, col).number_format = "0.0%"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{last_row}"


def add_options_sheet(wb, dropdown_options):
    ws = wb.create_sheet("選項")
    columns = [
        ("鞋款", dropdown_options["shoes"]),
        ("課表類型", dropdown_options["workout_types"]),
        ("訓練目的", dropdown_options["training_focus"]),
        ("感受難度", dropdown_options["garmin_rpe"]),
        ("感覺如何", dropdown_options["garmin_feel"]),
    ]
    for col, (title, options) in enumerate(columns, start=1):
        ws.cell(1, col, title)
        ws.cell(1, col).font = Font(name="Arial", bold=True)
        for row, option in enumerate(options, start=2):
            ws.cell(row, col, option)
        ws.column_dimensions[get_column_letter(col)].width = max(12, max(len(x) for x in [title, *options]) + 2)
    ws.sheet_state = "hidden"
    return ws


def add_metadata_sheet(wb, metadata, fit_path, session, rows, dropdown_options):
    metadata = coerce_metadata(metadata)
    if metadata.get("rpe", "") == "":
        metadata["rpe"] = garmin_rpe_label(session.get("workout_rpe"), dropdown_options["garmin_rpe"])
    else:
        metadata["rpe"] = normalize_rpe(metadata.get("rpe"), dropdown_options["garmin_rpe"])
    if metadata.get("feel", "") == "":
        metadata["feel"] = garmin_feel_label(session.get("workout_feel"), dropdown_options["garmin_feel"])
    else:
        metadata["feel"] = normalize_feel(metadata.get("feel"), dropdown_options["garmin_feel"])
    if metadata.get("training_effect_aerobic", "") == "":
        value = session.get("total_training_effect")
        if isinstance(value, (int, float)):
            metadata["training_effect_aerobic"] = round(float(value), 1)
    if metadata.get("training_effect_anaerobic", "") == "":
        value = session.get("total_anaerobic_training_effect")
        if isinstance(value, (int, float)):
            metadata["training_effect_anaerobic"] = round(float(value), 1)
    if metadata.get("training_load", "") == "":
        value = session.get("training_load_peak")
        if isinstance(value, (int, float)):
            metadata["training_load"] = round(float(value))

    start = fit_datetime(session.get("start_time") or session.get("timestamp"))
    activity = activity_summary(rows)
    economy = running_economy_summary(rows)
    stamina_start, stamina_end = stamina_summary(rows)
    metadata_sections = [
        (
            "Metadata",
            "1F4E78",
            "D9EAF7",
            [
                ("Excel Schema Version", "v1.1", "excel_schema_version"),
                ("資料來源", fit_path.name, "source"),
                ("Garmin Activity ID", metadata.get("garmin_activity_id", ""), "garmin_activity_id"),
                ("FIT Hash (SHA-256)", metadata.get("fit_sha256", ""), "fit_sha256"),
            ],
        ),
        (
            "Activity",
            "2E7D32",
            "E2F0D9",
            [
                ("活動日期", activity_date(session, fit_path), "activity_date"),
                ("開始時間", start.astimezone().strftime("%H:%M:%S") if start else "", "start_time"),
                ("活動類型", activity_type(session), "activity_type"),
                ("活動名稱", activity_name(session, metadata), "activity_name"),
                ("距離 (km)", activity["distance_km"], "distance_km"),
                ("時間", activity["duration"], "duration"),
                ("平均配速", activity["avg_pace"], "avg_pace"),
                ("課表類型", metadata.get("workout_type", ""), "workout_type"),
                ("訓練目的", metadata.get("training_focus", ""), "training_focus"),
                ("鞋款", metadata.get("shoe", ""), "shoe"),
            ],
        ),
        (
            "Environment",
            "B8860B",
            "FFF2CC",
            [
                ("天氣氣溫 (°C)", metadata.get("weather_temp", ""), "weather_temp"),
                ("濕度 (%)", metadata.get("humidity", ""), "humidity"),
                ("風向", metadata.get("wind_direction", ""), "wind_direction"),
                ("風速", metadata.get("wind_speed", ""), "wind_speed"),
                ("天氣描述", metadata.get("weather_description", ""), "weather_description"),
            ],
        ),
        (
            "Subjective",
            "C55A11",
            "FCE4D6",
            [
                ("感覺如何", metadata.get("feel", ""), "feel"),
                ("感受難度", metadata.get("rpe", ""), "rpe"),
                ("補給紀錄", metadata.get("fueling", ""), "fueling"),
                ("備註", metadata.get("notes", ""), "notes"),
            ],
        ),
        (
            "Training Metrics",
            "7030A0",
            "EADCF8",
            [
                ("最大心率", metadata.get("max_hr", ""), "max_hr"),
                ("Critical Power (W)", metadata.get("critical_power", ""), "critical_power"),
                ("Training Effect (Aerobic)", metadata.get("training_effect_aerobic", ""), "training_effect_aerobic"),
                ("Training Effect (Anaerobic)", metadata.get("training_effect_anaerobic", ""), "training_effect_anaerobic"),
                ("Training Load", metadata.get("training_load", ""), "training_load"),
                ("Recovery Time (hr)", metadata.get("recovery_time_hr", ""), "recovery_time_hr"),
                ("Stamina 起始 (%)", stamina_start, "stamina_start"),
                ("Stamina 結束 (%)", stamina_end, "stamina_end"),
            ],
        ),
        (
            "Running Economy",
            "156082",
            "DDEBF7",
            [
                ("平均步頻", economy["avg_cadence"], "avg_cadence"),
                ("平均步幅 (mm)", economy["avg_step_length"], "avg_step_length"),
                ("平均觸地時間 GCT (ms)", economy["avg_gct"], "avg_gct"),
                ("平均垂直振幅 (mm)", economy["avg_vertical_oscillation"], "avg_vertical_oscillation"),
                ("平均垂直比", economy["avg_vertical_ratio"], "avg_vertical_ratio"),
            ],
        ),
    ]

    ws = wb.create_sheet("活動資訊", 0)
    ws["A1"] = "活動資訊"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:B1")

    row_by_key = {}
    section_rows = {}
    row_fills = {}
    current_row = 2
    for section_name, header_color, fill_color, rows in metadata_sections:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(current_row, 1, section_name)
        ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=header_color)
        ws.cell(current_row, 1).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        ws.cell(current_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        section_rows[current_row] = header_color
        current_row += 1
        for label, value, key in rows:
            ws.cell(current_row, 1, label)
            ws.cell(current_row, 2, cell_value(value))
            row_by_key[key] = current_row
            row_fills[current_row] = fill_color
            current_row += 1

    option_ws = add_options_sheet(wb, dropdown_options)
    validations = {
        row_by_key["shoe"]: f"='選項'!$A$2:$A${len(dropdown_options['shoes']) + 1}",
        row_by_key["workout_type"]: f"='選項'!$B$2:$B${len(dropdown_options['workout_types']) + 1}",
        row_by_key["training_focus"]: f"='選項'!$C$2:$C${len(dropdown_options['training_focus']) + 1}",
        row_by_key["rpe"]: f"='選項'!$D$2:$D${len(dropdown_options['garmin_rpe']) + 1}",
        row_by_key["feel"]: f"='選項'!$E$2:$E${len(dropdown_options['garmin_feel']) + 1}",
    }
    for row, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row, 2))

    for row in (row_by_key["max_hr"], row_by_key["critical_power"]):
        positive_number = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1="0",
            allow_blank=True,
        )
        ws.add_data_validation(positive_number)
        positive_number.add(ws.cell(row, 2))

    for row in (
        row_by_key["training_effect_aerobic"],
        row_by_key["training_effect_anaerobic"],
        row_by_key["training_load"],
        row_by_key["recovery_time_hr"],
    ):
        non_negative_number = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        ws.add_data_validation(non_negative_number)
        non_negative_number.add(ws.cell(row, 2))

    thin_gray = Side(style="thin", color="D9E2F3")
    for row, fill_color in row_fills.items():
        fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row, 1).fill = fill
        ws.cell(row, 2).fill = fill
        ws.cell(row, 1).font = Font(name="Arial", bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).alignment = Alignment(horizontal="left")
        ws.cell(row, 1).border = Border(bottom=thin_gray)
        ws.cell(row, 2).border = Border(bottom=thin_gray)
    for row in section_rows:
        ws.cell(row, 1).border = Border(top=thin_gray, bottom=thin_gray)
        ws.cell(row, 2).border = Border(top=thin_gray, bottom=thin_gray)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72
    ws.freeze_panes = "A2"
    return option_ws


def add_percentage_values(ws, row_count, metadata):
    max_hr = metadata.get("max_hr")
    critical_power = metadata.get("critical_power")
    for row in range(3, row_count + 3):
        if isinstance(max_hr, (int, float)) and max_hr > 0 and isinstance(ws.cell(row, 5).value, (int, float)):
            ws.cell(row, 6, ws.cell(row, 5).value / max_hr)
        if (
            isinstance(critical_power, (int, float))
            and critical_power > 0
            and isinstance(ws.cell(row, 9).value, (int, float))
        ):
            ws.cell(row, 10, ws.cell(row, 9).value / critical_power)


def add_total_row(ws, row_count, rows):
    total_row = row_count + 3
    data_start = 3
    data_end = row_count + 2
    total_distance = sum(row[1] for row in rows if isinstance(row[1], (int, float)))
    total_seconds = sum(row[2] for row in rows if isinstance(row[2], (int, float)))
    ws.cell(total_row, 1, "總計/平均")
    ws.cell(total_row, 2, f"=SUM(B{data_start}:B{data_end})")
    ws.cell(total_row, 3, f"=SUM(C{data_start}:C{data_end})")
    ws.cell(total_row, 4, pace_text(total_seconds, total_distance))
    for col in range(5, 16):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=AVERAGE({letter}{data_start}:{letter}{data_end})")
    ws.cell(total_row, 16, f"=INDEX(P{data_start}:P{data_end},1)")
    ws.cell(total_row, 17, f"=INDEX(Q{data_start}:Q{data_end},{row_count})")
    ws.cell(total_row, 18, f"=SUM(R{data_start}:R{data_end})")
    return total_row


def add_charts(wb, row_count):
    ws = wb["每公里數據"]
    chart_ws = wb.create_sheet("圖表")
    chart_ws["A1"] = "配速 / 心率 / Stamina 趨勢圖"
    chart_ws["A1"].font = Font(name="Arial", size=14, bold=True)

    max_row = row_count + 2
    cats = Reference(ws, min_col=1, min_row=3, max_row=max_row)

    chart1 = LineChart()
    chart1.title = "心率與 Stamina 隨公里數變化"
    chart1.y_axis.title = "心率 / Stamina"
    chart1.x_axis.title = "公里"
    for col in (5, 16, 17):
        data = Reference(ws, min_col=col, min_row=2, max_row=max_row)
        chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 12
    chart1.width = 24
    chart_ws.add_chart(chart1, "A3")

    chart2 = LineChart()
    chart2.title = "配速趨勢(秒/公里，數值越低代表越快)"
    chart2.y_axis.title = "秒/公里"
    chart2.x_axis.title = "公里"
    data = Reference(ws, min_col=3, min_row=2, max_row=max_row)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 12
    chart2.width = 24
    chart_ws.add_chart(chart2, "A28")


def create_workbook(fit_path: Path, output_path: Path, metadata=None, fetch_weather=True, dropdown_options=None):
    dropdown_options = dropdown_options or load_dropdown_options()
    messages = decode_fit(fit_path)
    rows, session = build_rows(messages)
    if not rows:
        raise RuntimeError("No lap data found in FIT file.")
    metadata = apply_file_identity(metadata or {}, fit_path)
    metadata = apply_fit_metadata(metadata, messages)
    metadata = apply_auto_weather(metadata, session, messages.get("record_mesgs", []), fetch_weather)
    metadata = coerce_metadata(metadata)

    wb = Workbook()
    ws = wb.active
    ws.title = "每公里數據"
    add_metadata_sheet(wb, metadata, fit_path, session, rows, dropdown_options)
    date_label = activity_date(session, fit_path)
    ws["A1"] = f"{WORKBOOK_VERSION_NAME} - {date_label} (資料來源: {fit_path.name})"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    add_percentage_values(ws, len(rows), metadata)
    total_row = add_total_row(ws, len(rows), rows)
    apply_styles(ws, total_row)
    add_charts(wb, len(rows))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Convert a Garmin FIT running activity to a per-kilometer Excel workbook.")
    parser.add_argument("--version", action="version", version=f"Running Analytics v{APP_VERSION} / {WORKBOOK_VERSION_NAME}")
    parser.add_argument("fit_file", type=Path)
    parser.add_argument("-o", "--output", type=Path)
    parser.add_argument("--interactive", action="store_true", help="Prompt for manual activity metadata before exporting.")
    parser.set_defaults(fetch_weather=True)
    parser.add_argument("--fetch-weather", dest="fetch_weather", action="store_true", help="Fetch weather from Open-Meteo using FIT start time and GPS location. Enabled by default.")
    parser.add_argument("--no-fetch-weather", dest="fetch_weather", action="store_false", help="Skip automatic weather lookup.")
    parser.add_argument("--dropdown-config", type=Path, default=DROPDOWN_CONFIG_PATH, help="JSON file for dropdown options.")
    parser.add_argument("--activity-name", help="Activity name shown in the Activity section.")
    parser.add_argument("--shoe", help="Shoe name, e.g. 'Boston 13 Green'.")
    parser.add_argument("--weather-temp", type=float, help="Weather temperature in Celsius.")
    parser.add_argument("--humidity", type=float, help="Humidity percentage.")
    parser.add_argument("--wind-direction", help="Wind direction.")
    parser.add_argument("--wind-speed", help="Wind speed, e.g. '12 km/h'.")
    parser.add_argument("--weather-description", help="Weather description, e.g. sunny, cloudy, light rain.")
    parser.add_argument("--workout-type", help="Workout type, e.g. Recovery, Tempo, LSD, Intervals.")
    parser.add_argument("--training-focus", help="Training focus, e.g. Aerobic, Threshold, VO2max.")
    parser.add_argument("--feel", help="Workout feel, e.g. 50 or '普通'.")
    parser.add_argument("--rpe", help="Effort rating, e.g. 3, 30, or '3 - 中等'.")
    parser.add_argument("--fueling", help="Free-form fueling notes.")
    parser.add_argument("--max-hr", type=float, help="Maximum heart rate used for average heart rate percentage.")
    parser.add_argument("--critical-power", type=float, help="Critical Power in watts used for average power percentage.")
    parser.add_argument("--training-effect-aerobic", type=float, help="Aerobic Training Effect summary.")
    parser.add_argument("--training-effect-anaerobic", type=float, help="Anaerobic Training Effect summary.")
    parser.add_argument("--training-load", type=float, help="Training Load summary.")
    parser.add_argument("--recovery-time-hr", type=float, help="Recovery time in hours.")
    parser.add_argument("--notes", help="Free-form notes.")
    args = parser.parse_args()

    output = args.output
    if output is None:
        output = default_output_path(args.fit_file)

    dropdown_options = load_dropdown_options(args.dropdown_config)
    metadata = collect_metadata(args, dropdown_options)
    saved = create_workbook(
        args.fit_file,
        output,
        metadata,
        fetch_weather=args.fetch_weather,
        dropdown_options=dropdown_options,
    )
    print(saved)


if __name__ == "__main__":
    main()
